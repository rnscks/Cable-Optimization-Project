import os
from typing import List
from OCC.Core.gp import gp_Pnt
from OCC.Core.TopoDS import TopoDS_Shape
from openpyxl import load_workbook  

from src.brep.brep_util import STPFileReader
from src.grids_util import Voxelization
from src.grids.grids3d import Grids3D, Node
from src.cable import Cable


class Cabinet:
    def __init__(self) -> None:
        self.grids: Grids3D = None  
        self.cabinet_shape: TopoDS_Shape = None   
        self.cable_name_list: List[str] = self.get_cable_name_list()
        
        self.cabinet_shape = STPFileReader.read_stp_file_by_occ("CABINET.step")  
        self.grids = Grids3D(corner_max=gp_Pnt(200, 200, 200),
                corner_min=gp_Pnt(-200, -200, -200),
                map_size=30)
        Voxelization.voxelize(grids=self.grids, shape=self.cabinet_shape) 
        self.cables: List[Cable] = [self.get_cable(self.cable_name_list[0])]    
        
    
    def get_cable(self, cable_name: str) -> Cable:
        cable_data = self.get_cable_data(cable_name)
        start_pnt, end_pnt = cable_data['start_point'], cable_data['end_point']
        start_vec, end_vec = cable_data['start_vector'], cable_data['end_vector'] 
        middle_pnts = cable_data['middle_points']  
        middle_pnts = [gp_Pnt(*pnt) for pnt in middle_pnts]
        cable = Cable()

        cable.set_start_terminal(start_pnt=gp_Pnt(*start_pnt), start_vec = start_vec, grids=self.grids)
        cable.set_goal_terminal(goal_pos=gp_Pnt(*end_pnt), goal_vec = end_vec, grids=self.grids)
        cable.set_intermediate_pnts(pnts=middle_pnts, grids=self.grids, diameter=1)   

        #cable.set_spline(diameter=1, grids=self.grids)
        return cable

    def get_cable_data(self, cable_name):
        file_path = os.path.join(r'ModelCoordinates.xlsx')
        workbook = load_workbook(filename=file_path)
        
        # 각 시트 불러오기
        start_sheet = workbook['START']
        end_sheet = workbook['END']
        middle_sheet = workbook['MIDDLE']
        
        # 데이터를 저장할 사전
        cable_data = {
            'start_point': None,
            'start_vector': None,
            'end_point': None,
            'end_vector': None,
            'middle_points': []
        }
        
        # START 시트에서 데이터 추출
        for row in start_sheet.iter_rows(min_row=2):  # 첫 행은 헤더로 건너뛰기
            if row[0].value == cable_name:
                cable_data['start_point'] = (row[1].value, row[2].value, row[3].value)
                cable_data['start_vector'] = (row[4].value, row[5].value, row[6].value)
                break
        
        # END 시트에서 데이터 추출
        for row in end_sheet.iter_rows(min_row=2):
            if row[0].value == cable_name:
                cable_data['end_point'] = (row[1].value, row[2].value, row[3].value)
                cable_data['end_vector'] = (row[4].value, row[5].value, row[6].value)
                break
        
        # MIDDLE 시트에서 데이터 추출
        for row in middle_sheet.iter_rows(min_row=2):
            if row[0].value == cable_name:
                for i in range(1, len(row) - 2, 3):  # 3씩 증가하여 x, y, z 좌표 추출
                    # 빈칸 확인
                    if row[i].value is None or row[i+1].value is None or row[i+2].value is None:
                        break
                    cable_data['middle_points'].append((row[i].value, row[i+1].value, row[i+2].value))
        
        return cable_data
    
    def get_cable_name_list(self) -> List[str]:
        file_path = os.path.join(r'ModelCoordinates.xlsx')
        workbook = load_workbook(filename=file_path)
        
        # 각 시트 불러오기
        start_sheet = workbook['START']
        cable_name_list = []
        
        for row in start_sheet.iter_rows(min_row=2):
            if row[0].value is None:
                break
            cable_name_list.append(row[0].value)
            
        return cable_name_list 
    

            